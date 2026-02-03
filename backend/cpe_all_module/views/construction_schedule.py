from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..models.construction_schedule_models import ConstructionScheduleItem
from ..serializers.construction_schedule_serializers import ConstructionScheduleItemSerializer
from cpe_module.models.operating_rate_models import WorkScheduleWeight
from cpe_module.models.calc_models import WorkCondition
from cpe_module.models.project_models import Project

class ConstructionScheduleItemViewSet(viewsets.ModelViewSet):
    queryset = ConstructionScheduleItem.objects.all()
    serializer_class = ConstructionScheduleItemSerializer
    pagination_class = None

    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        project_id = self.request.query_params.get('project_id')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        return queryset
        
    def update(self, request, *args, **kwargs):
        # Log update request for debugging
        print("Update Request Data keys:", request.data.keys())
        if 'data' in request.data:
            print("Items count:", len(request.data['data']))
        return super().update(request, *args, **kwargs)

    @action(detail=False, methods=['post'])
    def initialize_default(self, request):
        project_id = request.data.get('project_id')
        if not project_id:
            return Response({"error": "project_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Cleanup legacy multiple rows if they exist
        existing = ConstructionScheduleItem.objects.filter(project_id=project_id)
        if existing.count() > 1:
            existing.delete()
        
        # Get or Create Single Container
        container, created = ConstructionScheduleItem.objects.get_or_create(project_id=project_id)
        
        # If it exists but has no data, or we just created it -> Populate Defaults
        if not container.data:
            from cpe_all_module.initial_data import get_default_schedule_data
            container.data = get_default_schedule_data()
            container.save()
            return Response({"message": "Initialized default items"}, status=status.HTTP_201_CREATED)
            
        return Response({"message": "Already initialized"}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response({"error": "project_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        container = ConstructionScheduleItem.objects.filter(project_id=project_id).first()
        if not container or not container.data:
            return Response({"error": "schedule data not found"}, status=status.HTTP_404_NOT_FOUND)

        raw_data = container.data
        items = raw_data if isinstance(raw_data, list) else raw_data.get("items", [])

        if not isinstance(items, list):
            return Response({"error": "invalid schedule data"}, status=status.HTTP_400_BAD_REQUEST)

        project = Project.objects.filter(id=project_id).first()
        project_name = project.title if project else "프로젝트"

        rate_qs = WorkScheduleWeight.objects.filter(project_id=project_id).values("main_category", "operating_rate")
        rate_map = {row["main_category"]: row["operating_rate"] for row in rate_qs}

        grouped = {}
        ordered_categories = []
        for item in items:
            category = item.get("main_category") or "기타"
            if category not in grouped:
                grouped[category] = []
                ordered_categories.append(category)
            grouped[category].append(item)

        wb = Workbook()
        ws = wb.active
        ws.title = "공사기간 산정 기준"

        header = [
            "구분",
            "공정",
            "공종",
            "수량산출(개산)",
            "단위",
            "내역수량",
            "생산성",
            "투입조",
            "생산량/일",
            "작업기간(W.D)",
            "가동율",
            "Calender Day",
            "비고"
        ]

        header_cols = len(header)
        header_fill = PatternFill("solid", fgColor="D9D9D9")
        category_fill = PatternFill("solid", fgColor="FFF7C7")
        title_fill = PatternFill("solid", fgColor="FFFFFF")
        thin = Side(border_style="thin", color="9E9E9E")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        header_font = Font(bold=True)
        category_font = Font(bold=True)
        title_font = Font(bold=True, size=13)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")

        work_condition = WorkCondition.objects.filter(project_id=project_id).first()
        region = work_condition.region if work_condition and work_condition.region else ""

        rate_texts = []
        work_week_days = None
        for idx, category in enumerate(ordered_categories, start=1):
            rate_value = rate_map.get(category)
            if rate_value is None:
                continue
            if work_week_days is None:
                weight = WorkScheduleWeight.objects.filter(project_id=project_id, main_category=category).first()
                if weight:
                    work_week_days = weight.work_week_days
            try:
                rate_value = float(rate_value)
                rate_texts.append(f"Cal-{idx}({category}) - {rate_value:.1f}%")
            except (TypeError, ValueError):
                rate_texts.append(f"Cal-{idx}({category}) - {rate_value}")

        work_week_label = f"주{work_week_days}일" if work_week_days else "주6일"
        region_label = f"{region} / " if region else ""
        rate_summary = (
            f"가동율({region_label}{work_week_label}) : " + ", ".join(rate_texts)
            if rate_texts
            else f"가동율({region_label}{work_week_label}) : -"
        )

        ws.append(["공사기간 산정 기준"] + [""] * (header_cols - 1))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=header_cols)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = left
        for col in range(1, header_cols + 1):
            ws.cell(row=1, column=col).border = border

        ws.append([f"공사명 : {project_name}"] + [""] * (header_cols - 1))
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=header_cols)
        ws.cell(row=2, column=1).alignment = left
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=5).value = rate_summary
        ws.cell(row=2, column=5).alignment = left
        ws.cell(row=2, column=5).font = Font(color="C00000", bold=True)
        for col in range(1, header_cols + 1):
            ws.cell(row=2, column=col).border = border

        ws.append(header)
        for col in range(1, header_cols + 1):
            cell = ws.cell(row=3, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        row_idx = 4
        for category in ordered_categories:
            category_items = grouped[category]
            ws.append([category] + [""] * (len(header) - 1))
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(header))
            cat_cell = ws.cell(row=row_idx, column=1)
            cat_cell.fill = category_fill
            cat_cell.font = category_font
            cat_cell.alignment = left
            for col in range(1, len(header) + 1):
                ws.cell(row=row_idx, column=col).border = border
            row_idx += 1

            for item in category_items:
                rate_value = rate_map.get(item.get("main_category"))
                if rate_value is None:
                    rate_value = item.get("operating_rate_value", "")
                if rate_value not in ("", None):
                    try:
                        rate_value = float(rate_value)
                    except (TypeError, ValueError):
                        pass

                row = [
                    item.get("main_category", ""),
                    item.get("process", ""),
                    item.get("work_type", ""),
                    item.get("quantity_formula", ""),
                    item.get("unit", ""),
                    item.get("quantity", ""),
                    item.get("productivity", ""),
                    item.get("crew_size", ""),
                    item.get("daily_production", ""),
                    item.get("working_days", ""),
                    rate_value,
                    item.get("calendar_days", ""),
                    item.get("remarks", "")
                ]
                ws.append(row)
                for col_idx in range(1, len(header) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border
                    if col_idx in (6, 7, 8, 9, 10, 11, 12):
                        cell.alignment = right
                        if col_idx in (6, 7, 9, 10, 12):
                            cell.number_format = '#,##0.00'
                        if col_idx == 8:
                            cell.number_format = '#,##0.0'
                        if col_idx == 11 and rate_value != "":
                            cell.number_format = '0.0"%"'
                    else:
                        cell.alignment = left
                row_idx += 1

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 38
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 8
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 10
        ws.column_dimensions["L"].width = 14
        ws.column_dimensions["M"].width = 28

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 20

        safe_name = "".join(ch if ch not in '\\/:*?"<>|' else "_" for ch in project_name)
        filename = f"공사기간_산정_기준_{safe_name}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
