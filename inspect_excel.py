
import pandas as pd
from openpyxl import load_workbook

file_path = '/home/pengejeen/p6ix-CPE/backend/cpe_all_module/data/260105_생산성 분석.xlsx'

try:
    # 1. Read Values using pandas to get data
    df = pd.read_excel(file_path, sheet_name='CIP 생산성 근거', header=None)
    
    print("\n--- Rows 6-10 (Excel Rows 7-11) ---")
    # Columns:
    # O (14) = t1
    # P (15) ~ W (22)? Let's print a range.
    # Openpyxl: A=1. O=15.
    
    # Let's print values for Row 8 (Excel Row 9) which is the first data row
    print("Row 8 (Excel Row 9) Data:")
    print(df.iloc[8].to_string())

    # 2. Read Formulas using openpyxl
    wb = load_workbook(file_path, data_only=False)
    ws = wb['CIP 생산성 근거']
    
    print("\n--- Formulas in Row 9 ---")
    # t1 is likely Column O.
    # t2 Time is likely Column V (22) or W?
    # t3 is likely Column W (23)?
    # t4 Length is X (24)?
    # t4 Time is Y (25)?
    
    # Let's just scan columns K to Z for Row 9
    for col_idx in range(11, 27): # K is 11th letter
        col_letter =  chr(64 + col_idx) # Simple conversion for A-Z
        if col_idx > 26: continue # Skip AA etc for now if not needed
        
        cell_ref = f"{col_letter}9"
        cell = ws[cell_ref]
        print(f"Cell {cell_ref}: Value={cell.value}")
        
    # Also check the Standards Table at the bottom (Row 28+)
    # Column D(4) = Clay, E(5)=Sand, etc.
    print("\n--- Standards Table (Rows 28-33) ---")
    print(df.iloc[28:34, 1:10].to_string())

except Exception as e:
    print(f"Error: {e}")
